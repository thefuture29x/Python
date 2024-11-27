import warnings
import sys
from openpyxl import load_workbook

x_loop = 1
y_loop = 2

# Suppress openpyxl warning
warnings.simplefilter("ignore", UserWarning)

# Change the console encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

excel_file = load_workbook("P:\Excel\Oder_2022-07-01_2022-07-06.xlsx")
sheet = excel_file.active

# Access data from the sheet

while True:
    cell_data = sheet.cell(column=x_loop, row=y_loop).value
    if cell_data != '':
        print(str(y_loop) + "---->" +cell_data)
        y_loop += 1
    else:
        break
    
# Close the Excel file
excel_file.close()
